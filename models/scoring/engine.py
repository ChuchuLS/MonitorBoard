"""
models/scoring/engine.py
=======================
Global Rates & Equity cross-sectional scoring model.

Extracted from the Pulsar/CTA Dashboard. Pure math — no Streamlit.
Scores 10 sovereign bond markets and 17 equity indices on macro + market
factors, producing a cross-sectional z-score ranking.
"""

from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import pandas as pd

# UNIVERSES & MAPPINGS
# ============================================================

# Rates dashboard: 10 sovereign bond markets
RATES_UNIVERSE = [
    ("FR", "France"),
    ("JP", "Japan"),
    ("GB", "UK"),
    ("ES", "Spain"),
    ("IT", "Italy"),
    ("US", "United States"),
    ("CA", "Canada"),
    ("KR", "S. Korea"),
    ("AU", "Australia"),
    ("DE", "Germany"),
]
RATES_CODES = [c for c, _ in RATES_UNIVERSE]

# Equity dashboard: 17 indices
EQUITY_UNIVERSE = [
    ("PT1",  "S&P/TSX",        "Canada"),
    ("NQ1",  "Nasdaq 100",     "USA"),
    ("KM1",  "KOSPI",          "S. Korea"),
    ("XP1",  "ASX",            "Australia"),
    ("HI1",  "Hang Seng",      "Hong Kong"),
    ("XU1",  "FTSE China A50", "China"),
    ("ES1",  "S&P 500",        "USA"),
    ("RTY1", "Russell 2000",   "USA"),
    ("SM1",  "SMI",            "Switzerland"),
    ("Z 1",  "FTSE 100",       "UK"),
    ("EO1",  "AEX",            "Netherlands"),
    ("CF1",  "CAC 40",         "France"),
    ("ST1",  "FTSE MIB",       "Italy"),
    ("NK1",  "Nikkei 225",     "Japan"),
    ("VG1",  "Euro Stoxx 50",  "Eurozone"),
    ("IB1",  "IBEX 35",        "Spain"),
    ("GX1",  "DAX",            "Germany"),
]
EQUITY_CODES = [c for c, _, _ in EQUITY_UNIVERSE]
EQUITY_META = {c: (n, r) for c, n, r in EQUITY_UNIVERSE}

# Each equity index → its macro country (for GDP/CPI/Fiscal lookup)
INDEX_TO_COUNTRY = {
    "PT1": "CA", "NQ1": "US", "KM1": "KR", "XP1": "AU",
    "HI1": "HK", "XU1": "CN", "ES1": "US", "RTY1": "US",
    "SM1": "CH", "Z 1": "GB", "EO1": "NL", "CF1": "FR",
    "ST1": "IT", "NK1": "JP", "VG1": "EZ", "IB1": "ES", "GX1": "DE",
}

# Each equity index → its FCI region (we have 4 FCIs only)
# Goldman/Bloomberg only publish FCI for major financial systems
INDEX_TO_FCI_REGION = {
    "PT1": "NQ1",   # Canada → US FCI (closest financial system proxy)
    "NQ1": "NQ1",   # US
    "KM1": "NQ1",   # Korea → US FCI (USD-funded EM)
    "XP1": "NQ1",   # Australia → US FCI
    "HI1": "XU1",   # Hang Seng → China FCI
    "XU1": "XU1",   # China
    "ES1": "NQ1",   # US
    "RTY1": "NQ1",  # US
    "SM1": "VG1",   # Switzerland → Eurozone FCI
    "Z 1": "Z 1",   # UK
    "EO1": "VG1",   # Netherlands → Eurozone FCI
    "CF1": "VG1",   # France → Eurozone FCI
    "ST1": "VG1",   # Italy → Eurozone FCI
    "NK1": "NQ1",   # Japan → US FCI (open economy fallback)
    "VG1": "VG1",   # Eurozone
    "IB1": "VG1",   # Spain → Eurozone FCI
    "GX1": "VG1",   # Germany → Eurozone FCI
}

# Each equity index → its Citi ToT currency ticker
INDEX_TO_TOT_TICKER = {
    "PT1": "CTOTCAD Index", "NQ1": "CTOTUSD Index", "KM1": "CTOTKRW Index",
    "XP1": "CTOTAUD Index", "HI1": "CTOTHKD Index", "XU1": "CTOTCNY Index",
    "ES1": "CTOTUSD Index", "RTY1": "CTOTUSD Index", "SM1": "CTOTCHF Index",
    "Z 1": "CTOTGBP Index", "EO1": "CTOTEUR Index", "CF1": "CTOTEUR Index",
    "ST1": "CTOTEUR Index", "NK1": "CTOTJPY Index", "VG1": "CTOTEUR Index",
    "IB1": "CTOTEUR Index", "GX1": "CTOTEUR Index",
}

# Each rates country → its Citi ToT currency
RATES_TO_TOT_TICKER = {
    "FR": "CTOTEUR Index", "JP": "CTOTJPY Index", "GB": "CTOTGBP Index",
    "ES": "CTOTEUR Index", "IT": "CTOTEUR Index", "US": "CTOTUSD Index",
    "CA": "CTOTCAD Index", "KR": "CTOTKRW Index", "AU": "CTOTAUD Index",
    "DE": "CTOTEUR Index",
}


# ============================================================
# LOADER
# ============================================================
def read_sheet(xlsx_path: str, sheet_name: str, header_row: int = 4) -> pd.DataFrame:
    """
    Read a sheet with row 4 as code headers, dates in col A from row 6.
    Handles BDH-style 'Date' interleaved columns by skipping them.
    Returns DataFrame indexed by date with code columns.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    # Read row 4 codes; drop any column literally labeled "Date"
    raw_codes = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        raw_codes.append((c, str(v).strip() if v is not None else None))

    # Keep only columns whose code is not None and not "Date"
    valid_cols = [(c, code) for c, code in raw_codes if code and code != "Date"]
    codes = [code for _, code in valid_cols]
    col_idx = [c for c, _ in valid_cols]

    rows = []
    for r in range(6, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is None or (isinstance(d, str) and (d.strip() == "" or "#N/A" in d)):
            continue
        try:
            ts = pd.Timestamp(d)
        except Exception:
            continue
        vals = []
        for c in col_idx:
            v = ws.cell(row=r, column=c).value
            if v is None or v == "" or (isinstance(v, str) and "#N/A" in v):
                vals.append(np.nan)
            else:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    vals.append(np.nan)
        rows.append([ts] + vals)

    df = pd.DataFrame(rows, columns=["date"] + codes).set_index("date").sort_index()
    df.index = pd.to_datetime(df.index)
    df = df[~df.index.duplicated(keep="last")]
    return df


def load_all(xlsx_path: str) -> dict:
    return {
        "gdp":    read_sheet(xlsx_path, "Macro_GDP"),
        "cpi":    read_sheet(xlsx_path, "Macro_CPI"),
        "fiscal": read_sheet(xlsx_path, "Macro_Fiscal"),
        "y10y":   read_sheet(xlsx_path, "Rates_10Y"),
        "tot":    read_sheet(xlsx_path, "Equity_ToT"),
        "fci":    read_sheet(xlsx_path, "Equity_FCI"),
        "eps":    read_sheet(xlsx_path, "Equity_EPS"),
        "px":     read_sheet(xlsx_path, "Equity_Prices"),
    }


# ============================================================
# FACTOR MATH HELPERS
# ============================================================
def latest(df: pd.DataFrame, asof: pd.Timestamp) -> pd.Series:
    """Latest available value per column on or before asof."""
    sub = df.loc[:asof].ffill()
    if len(sub) == 0:
        return pd.Series(dtype=float)
    return sub.iloc[-1]


def value_n_days_ago(df: pd.DataFrame, asof: pd.Timestamp, days: int) -> pd.Series:
    sub = df.loc[:asof].ffill()
    if len(sub) == 0:
        return pd.Series(dtype=float)
    target = asof - pd.Timedelta(days=days)
    if len(sub.loc[:target]) > 0:
        return sub.loc[:target].iloc[-1]
    return sub.iloc[0]


def pct_change(df: pd.DataFrame, asof: pd.Timestamp, days: int) -> pd.Series:
    last = latest(df, asof)
    base = value_n_days_ago(df, asof, days)
    return (last / base - 1) * 100


def diff_n_days(df: pd.DataFrame, asof: pd.Timestamp, days: int) -> pd.Series:
    """Absolute (not %) change over n days. For yields, ToT etc."""
    last = latest(df, asof)
    base = value_n_days_ago(df, asof, days)
    return last - base


def realized_vol(df: pd.DataFrame, asof: pd.Timestamp, window: int = 30) -> pd.Series:
    """Annualized rolling std of daily log returns."""
    sub = df.loc[:asof].ffill()
    if len(sub) < 2:
        return pd.Series(index=df.columns, dtype=float)
    rets = np.log(sub).diff()
    tail = rets.tail(window)
    return tail.std() * np.sqrt(252) * 100


def zscore(series: pd.Series, codes: list, sign: int = 1) -> pd.Series:
    """Cross-sectional z-score over the given universe; sign flips direction."""
    s = series.reindex(codes).astype(float)
    mu = s.mean(skipna=True)
    sd = s.std(ddof=0, skipna=True)
    if sd == 0 or np.isnan(sd):
        return s * 0
    return sign * (s - mu) / sd


# ============================================================
# RATES SCORING
# ============================================================
def score_rates(data: dict, asof: pd.Timestamp, weights: dict) -> pd.DataFrame:
    """
    Replicate the Pulsar Rates dashboard:
    Macro pillar = mean(GDP_z, CPI_z (inverted), Budget_z)
    Markets pillar = mean(Momentum_z, Carry_z, RealYield_z)
    Score = mean(Macro, Markets)
    """
    gdp_v    = latest(data["gdp"],    asof)
    cpi_v    = latest(data["cpi"],    asof)
    fiscal_v = latest(data["fiscal"], asof)
    y10y_v   = latest(data["y10y"],   asof)
    y10y_3m  = value_n_days_ago(data["y10y"], asof, 90)

    # Build factor frames per country (only rates universe)
    z_gdp    = zscore(gdp_v,    RATES_CODES, +1)   # higher GDP = good for govt
    z_cpi    = zscore(cpi_v,    RATES_CODES, -1)   # lower inflation = good for bonds
    z_budget = zscore(fiscal_v, RATES_CODES, +1)   # less deficit (less negative) = good
    macro    = pd.concat([z_gdp, z_cpi, z_budget], axis=1).mean(axis=1)

    # 3M yield change: bonds rally when yields FALL → lower change = positive
    z_mom    = zscore(y10y_v - y10y_3m, RATES_CODES, -1)
    z_carry  = zscore(y10y_v,            RATES_CODES, +1)   # higher carry = good
    # Real yield = nominal − CPI
    real_y   = y10y_v - cpi_v
    z_realy  = zscore(real_y,            RATES_CODES, +1)
    markets  = pd.concat([z_mom, z_carry, z_realy], axis=1).mean(axis=1)

    # Weighted composite
    wsum = sum(weights.values())
    wn = {k: v / wsum for k, v in weights.items()}
    score = macro * wn["macro"] + markets * wn["markets"]

    out = pd.DataFrame({
        "gdp_z":    z_gdp,
        "cpi_z":    z_cpi,
        "budget_z": z_budget,
        "macro":    macro,
        "mom_z":    z_mom,
        "carry_z":  z_carry,
        "realy_z":  z_realy,
        "markets":  markets,
        "score":    score,
    }).round(2)
    out.index.name = "code"
    out["country"] = [dict(RATES_UNIVERSE)[c] for c in out.index]
    out["incomplete"] = pd.concat([macro, markets], axis=1).isna().any(axis=1)
    return out.sort_values("score", ascending=False, na_position="last")


# ============================================================
# EQUITY SCORING
# ============================================================
def score_equity(data: dict, asof: pd.Timestamp, weights: dict) -> pd.DataFrame:
    """
    Pulsar Equity dashboard:
    Macro pillar = mean(Growth_z, Inflation_z (inv), Deficit_z, ToT_z, FCI_z)
    Then composite score = weighted blend of macro + EPS revisions.
    Plus performance and vol columns.
    """
    gdp_v    = latest(data["gdp"],    asof)
    cpi_v    = latest(data["cpi"],    asof)
    fiscal_v = latest(data["fiscal"], asof)
    tot_v    = latest(data["tot"],    asof)
    tot_3m   = value_n_days_ago(data["tot"], asof, 90)
    fci_v    = latest(data["fci"],    asof)

    # Map each index to its country macro values
    def by_index(country_series: pd.Series) -> pd.Series:
        out = {}
        for code in EQUITY_CODES:
            country = INDEX_TO_COUNTRY[code]
            out[code] = country_series.get(country, np.nan)
        return pd.Series(out)

    growth     = by_index(gdp_v)
    inflation  = by_index(cpi_v)
    deficit    = by_index(fiscal_v)

    # ToT: 3M change in the index's currency ToT
    def tot_for_index(asof_series, three_m_series):
        out_now, out_then = {}, {}
        for code in EQUITY_CODES:
            t = INDEX_TO_TOT_TICKER.get(code)
            out_now[code]  = asof_series.get(t, np.nan)
            out_then[code] = three_m_series.get(t, np.nan)
        return pd.Series(out_now), pd.Series(out_then)
    tot_now, tot_then = tot_for_index(tot_v, tot_3m)
    tot_mom = tot_now - tot_then  # absolute change in ToT index

    # FCI: latest value of the index's regional FCI
    fci_by_index = pd.Series({code: fci_v.get(INDEX_TO_FCI_REGION[code], np.nan)
                              for code in EQUITY_CODES})

    # EPS Δ: 3M % change in FY1 EPS estimate
    eps_v   = latest(data["eps"], asof)
    eps_3m  = value_n_days_ago(data["eps"], asof, 90)
    eps_chg = (eps_v / eps_3m - 1) * 100
    # Align to equity universe (eps frame columns are already index codes)
    eps_delta = eps_chg.reindex(EQUITY_CODES)

    # Z-scores on equity universe
    z_growth = zscore(growth,    EQUITY_CODES, +1)
    z_infl   = zscore(inflation, EQUITY_CODES, -1)
    z_def    = zscore(deficit,   EQUITY_CODES, +1)
    z_tot    = zscore(tot_mom,   EQUITY_CODES, +1)
    z_fci    = zscore(fci_by_index, EQUITY_CODES, +1)

    macro = pd.concat([z_growth, z_infl, z_def, z_tot, z_fci], axis=1).mean(axis=1)

    # EPS as separate factor (z-scored across the panel)
    z_eps = zscore(eps_delta, EQUITY_CODES, +1)

    # Weighted composite
    wsum = sum(weights.values())
    wn = {k: v / wsum for k, v in weights.items()}
    score = macro * wn["macro"] + z_eps * wn["eps"]

    # Performance and vol (cosmetic columns)
    p5d = pct_change(data["px"], asof, 7).reindex(EQUITY_CODES)
    p1m = pct_change(data["px"], asof, 30).reindex(EQUITY_CODES)
    p3m = pct_change(data["px"], asof, 90).reindex(EQUITY_CODES)
    vol = realized_vol(data["px"], asof, 30).reindex(EQUITY_CODES)

    out = pd.DataFrame({
        "growth_z": z_growth,
        "infl_z":   z_infl,
        "def_z":    z_def,
        "tot_z":    z_tot,
        "fci_z":    z_fci,
        "macro":    macro,
        "eps_delta": eps_delta,
        "score":    score,
        "p5d": p5d,
        "p1m": p1m,
        "p3m": p3m,
        "vol": vol,
    }).round(2)
    out.index.name = "code"
    out["name"]   = [EQUITY_META[c][0] for c in out.index]
    out["region"] = [EQUITY_META[c][1] for c in out.index]
    out["incomplete"] = macro.isna() | z_eps.isna()
    return out.sort_values("score", ascending=False, na_position="last")


# ============================================================
# RENDERING (Pulsar HTML style)
# ============================================================
