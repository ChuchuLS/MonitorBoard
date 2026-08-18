"""Optional export integration test — run separately from routine smoke tests.

Usage:
    python smoke_test_export.py
    FULL_EXPORT_SMOKE=1 python smoke_test_export.py
"""
import sys, os, time, subprocess
from pathlib import Path
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
_t0 = time.time()

print("=== EXPORT INTEGRATION TEST ===\n")

# 1. Snapshot command-line entry
print("1. Snapshot CLI entry point ...")
result = subprocess.run(["python", "scripts/export_research_pack_snapshot.py"],
                        capture_output=True, text=True, timeout=120)
assert result.returncode == 0, f"snapshot script failed: {result.stderr[:400]}"
assert Path("data/snapshot.json").exists()
import json
snap_data = json.loads(Path("data/snapshot.json").read_text())
assert "index" in snap_data and "pages" in snap_data
assert snap_data["index"].get("status") == "Official"
assert snap_data["index"].get("as_of"), "snapshot official as-of date missing"
assert snap_data.get("export_errors") == [], \
    f"snapshot sections failed: {snap_data.get('export_errors')}"
print(f"   snapshot.json: index={snap_data['index']['level']}, {len(snap_data['pages'])} pages ✓")

# 2. Multi-sheet Excel export
print("\n2. Multi-sheet Excel export ...")
from data.loader import load_data, source_signature
from index.composite import compute_index
from index.export import build_index_workbook
from index.methodology import (
    component_contribution_table, compute_legacy_index, forward_fill_audit,
    methodology_audit, reconciliation,
)
frame = load_data()
index_result = compute_index(frame)
legacy_result = compute_legacy_index(frame)
audit = {
    "methodology": methodology_audit(index_result, frame, data_hash=source_signature()),
    "reconciliation": reconciliation(index_result, legacy_result, frame),
    "components": component_contribution_table(index_result, frame),
    "ffill_audit": forward_fill_audit(index_result, frame),
}
xlsx_bytes = build_index_workbook(index_result, audit, frame)
assert xlsx_bytes.startswith(b"PK"), "Excel workbook ZIP signature missing"
import openpyxl
workbook = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True)
expected_sheets = {
    "Index", "Buckets", "Component_z", "Component_contrib",
    "Latest_components", "Reconciliation", "Forward_fill_audit", "Methodology",
}
assert expected_sheets.issubset(workbook.sheetnames), workbook.sheetnames
index_headers = [cell.value for cell in next(workbook["Index"].iter_rows(min_row=1, max_row=1))]
for required in ("index", "analytical_index", "preliminary_index",
                 "normal_component_target", "complete_coverage", "published"):
    assert required in index_headers, f"Index export missing {required}"
print(f"   Excel: {len(xlsx_bytes):,} bytes, {len(workbook.sheetnames)} sheets ✓")

# 3. Lightweight HTML export (plotly_mode=none)
print("\n3. Lightweight HTML export (plotly_mode=none) ...")
t_light = time.time()
from scripts.export_research_pack_html import build_html
html_str, filename = build_html(include_plotlyjs=False, plotly_mode="none")
elapsed_light = time.time() - t_light
assert len(html_str) > 1000
assert "Plotly.newPlot" not in html_str, "lightweight mode must not embed Plotly JS"
assert "<script>" not in html_str.split("</head>")[0], "lightweight mode must not include Plotly JS"
print(f"   lightweight: {len(html_str):,} chars in {elapsed_light:.2f}s ✓")

# 4. Complete PDF export
print("\n4. Complete linked PDF export ...")
from scripts.export_research_pack_pdf import build_pdf
pdf_bytes, pdf_name = build_pdf()
assert pdf_bytes.startswith(b"%PDF-"), "PDF signature missing"
assert len(pdf_bytes) > 50_000, "PDF export is unexpectedly small"
assert pdf_name.startswith("rates_liquidity_board_") and pdf_name.endswith(".pdf")
try:
    from pypdf import PdfReader
    reader = PdfReader(BytesIO(pdf_bytes))
    assert len(reader.pages) == 2 + len(snap_data["pages"]), \
        "PDF must contain cover + contents + every registered Board page"
    assert reader.outline, "PDF bookmarks are missing"
    print(f"   PDF: {len(reader.pages)} pages, {len(pdf_bytes):,} bytes, bookmarks present ✓")
except ImportError:
    print(f"   PDF: {len(pdf_bytes):,} bytes ✓ (pypdf structure check skipped)")

# 5. CTA CSV/HTML report export
print("\n5. CTA backtest CSV/HTML export ...")
cta_result = subprocess.run(
    [sys.executable, "scripts/run_cta_score_backtest.py"],
    capture_output=True, text=True, timeout=120,
)
assert cta_result.returncode == 0, cta_result.stderr[:500]
cta_paths = [Path(line.strip()) for line in cta_result.stdout.splitlines() if line.strip()]
assert len(cta_paths) == 3 and all(path.exists() for path in cta_paths), cta_paths
import pandas as pd
equity_csv = next(path for path in cta_paths if "equity" in path.name)
rates_csv = next(path for path in cta_paths if "rates" in path.name)
assert len(pd.read_csv(equity_csv)) == len(snap_data["cta_score_backtest"]["equity_periods"])
assert len(pd.read_csv(rates_csv)) == len(snap_data["cta_score_backtest"]["rates_periods"])
print(f"   CTA: {equity_csv.name} + {rates_csv.name} + HTML ✓")

# 6. Full inline Plotly export (opt-in via env var)
if os.environ.get("FULL_EXPORT_SMOKE") == "1":
    print("\n6. FULL inline Plotly HTML export ...")
    # Run the heavy inline build in a fresh process. Reusing the process after
    # lightweight mode can leave Plotly/module-level render state in a slow or
    # non-deterministic state in some headless environments.
    t_full = time.time()
    full_code = (
        "from scripts.export_research_pack_html import build_html; "
        "h, f = build_html(include_plotlyjs=True, plotly_mode='inline'); "
        "assert len(h) > 1000000 and 'Plotly.newPlot' in h; "
        "print(len(h))"
    )
    full_result = subprocess.run(
        ["python", "-c", full_code], capture_output=True, text=True, timeout=120
    )
    assert full_result.returncode == 0, full_result.stderr[:500]
    full_len = int(full_result.stdout.strip().splitlines()[-1])
    elapsed_full = time.time() - t_full
    print(f"   full: {full_len:,} chars in {elapsed_full:.2f}s ✓")
else:
    print("\n6. FULL inline Plotly test SKIPPED (set FULL_EXPORT_SMOKE=1 to enable)")

print(f"\nALL EXPORT TESTS PASSED ✓")
print(f"Total elapsed: {time.time() - _t0:.2f}s")
